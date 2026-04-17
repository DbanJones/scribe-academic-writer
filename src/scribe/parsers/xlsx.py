"""XLSX text extraction using openpyxl."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def extract(path: Path, sheet_names: list[str] | None = None) -> str:
    """Extract spreadsheet data as markdown tables.

    Args:
        path: Path to the XLSX file.
        sheet_names: If given, only extract these sheets. Otherwise all sheets.
    """
    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []

    sheets = sheet_names if sheet_names else wb.sheetnames

    for name in sheets:
        if name not in wb.sheetnames:
            parts.append(f"## Sheet: {name}\n(Sheet not found)")
            continue

        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        parts.append(f"## Sheet: {name}")

        # Filter out completely empty rows
        rows = [r for r in rows if any(c is not None for c in rows[0])]
        if not rows:
            continue

        # Header
        header = [str(c) if c is not None else "" for c in rows[0]]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")

        # Data rows (cap at 500 to avoid enormous outputs)
        for row in rows[1:501]:
            cells = [str(c) if c is not None else "" for c in row]
            parts.append("| " + " | ".join(cells) + " |")

        if len(rows) > 501:
            parts.append(f"\n(... {len(rows) - 501} more rows truncated)")

        parts.append("")

    wb.close()
    return "\n\n".join(parts)
